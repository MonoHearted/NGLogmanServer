import grpc
import sys
import os
import uuid
import re
from openpyxl import Workbook, load_workbook
from time import sleep

from . import nglm_pb2
from . import nglm_pb2_grpc
from nglogman.models import LGNode, Task, NodeGroup
from nglm_grpc.modules.Utility import timestamp
from apscheduler.schedulers.background import BackgroundScheduler

"""
Contains the methods that relate to the gRPC methods used in NGLogman. This
contain procedures called by the client to the server, and vice versa.
"""

SCHEDULER = BackgroundScheduler()
TIMEOUT_SECS = 2
ROOT_DIR = os.path.dirname(sys.modules['__main__'].__file__)

class ServerServicer(nglm_pb2_grpc.ServerServicer):
    def register(self, request, context):
        """
        Register procedure called by client. Should not be called manually.
        :param request: Attributes are hostname, ipv4, port, uuid. Contains
        the information sent by the client in the registration call.
        :param context: gRPC context object. Unused in this function.
        :return: A boolean to the client describing success.
        """
        res = nglm_pb2.registerResponse()
        try:
            matchedNodes = LGNode.objects.filter(
                hostname__iexact=request.hostname,
                ip__iexact=request.ipv4
            )
            if matchedNodes.count() == 0:
                res.uuid = request.uuid if request.uuid else str(uuid.uuid4())
                LGNode.objects.create(hostname=request.hostname,
                                      ip=request.ipv4, port=request.port,
                                      nodeUUID=uuid.UUID(res.uuid))
                print('Registered new node. Hostname: ' +
                      request.hostname + ' IP: ' + request.ipv4 +
                      ' Host Port:' + str(request.port))
            else:
                matchedNodes.update(port=request.port)
                res.uuid = str(matchedNodes[0].nodeUUID)
                print('Returning node. Hostname: ' +
                      request.hostname + ' IP: ' + request.ipv4 +
                      ' Host Port:' + str(request.port))
            res.success = True
        except Exception as e:
            res.success = False
            raise e
        return res

    def isAlive(self, request, context):
        """
        A simple receiver for a client-to-server healthcheck.
        :param request: gRPC request object. Unused in this function.
        :param context: gRPC context object. Unused in this function.
        :return: Always returns True (A failed check would not reach.)
        """
        res = nglm_pb2.response()
        res.success = True
        return res


class LoggingServicer(nglm_pb2_grpc.LoggingServicer):
    def output(self, request, context):
        """
        Receives a bytestream of the excel file returned by clients at the end
        of task execution. Saves the output file, and checks for completion.
        :param request: A byte stream representing the output xlsx from client.
        :param context: Metadata in context provides the UUID of the client.
        :return: A boolean representing whether the call was successful.
        """
        res = nglm_pb2.response()
        try:
            metadata = context.invocation_metadata()
            node_uuid = uuid.UUID(metadata[0].value)
            node = LGNode.objects.filter(nodeUUID=node_uuid)[0]
            task = Task.objects.filter(taskUUID=node.currentTask)[0]

            saveResponse(request, os.path.join(
                ROOT_DIR, "Reports", task.taskName + '_' + str(task.taskUUID),
                timestamp(task.startTime),
                node.hostname + '_' + node.ip.split('.')[-1] + '_' +
                str(task.taskUUID) + '_result.xlsx'))
            LGNode.objects.filter(nodeUUID=node.nodeUUID).update(
                status="Available", currentTask=None)
            res.success = True
            validateTask(task)
        except:
            res.success = False
        return res

    def err(self, request, context):
        """
        Called by the client should an error be raised during execution.
        Updates the status of the task to represent the failure.
        :param request: Contains the error message raised by the client.
        :param context: Metadata contains the UUID of the client.
        :return: Boolean representing error receiving success.
        """
        res = nglm_pb2.response()
        try:
            metadata = context.invocation_metadata()
            node_uuid = uuid.UUID(metadata[0].value)
            node = LGNode.objects.filter(nodeUUID=node_uuid)[0]
            task = Task.objects.filter(taskUUID=node.currentTask)[0]

            if 'Failed' in task.status:
                task.status = task.status + '\n%s: %s' \
                              % (node.ip, request.exception)
            else:
                task.status = 'Failed: \n%s: %s' \
                            % (node.ip, request.exception)
            task.save()
            NodeGroup.objects.filter(currentTask=task.taskUUID).update(
                currentTask=None)
            node.status = 'Available'
            node.currentTask = None
            node.save()

            res.success = True
        except:
            res.success = False
        return res


def addToServer(server):
    # Adds services to server, called on server start
    """
    Registers the gRPC Services on the server. Should new services be
    added, add their Servicer below.
    :param server: The grpc.server instance to add the servicer to.
    :return: None.
    """
    nglm_pb2_grpc.add_ServerServicer_to_server(ServerServicer(), server)
    nglm_pb2_grpc.add_LoggingServicer_to_server(LoggingServicer(), server)


def validateTask(task):
    """
    Called when the server receives an output, and checks for task completion.
    If the task is completed, generate the overview workbook, and update task
    status.
    :param task: The task to check completion for.
    :return: A boolean representing whether the task is completed or not.
    """
    output_path = os.path.join(ROOT_DIR, "Reports",
                               task.taskName + '_' + str(task.taskUUID),
                               timestamp(task.startTime))
    for node in task.assignedNode.nodes.all():
        path = os.path.join(
            output_path,
            node.hostname + '_' + node.ip.split('.')[-1] + '_' +
            str(task.taskUUID) + '_result.xlsx'
        )
        if not os.path.isfile(path):
            return False

    res_path = os.path.join(output_path, 'overview.xlsx')
    wb = Workbook()
    for root, dirs, files in os.walk(output_path):
        for f in files:
            name = re.compile(r'(.*)_.*_(?:result.xlsx)').match(f).group(1)
            wb.create_sheet(name, 0)
            read_wb = load_workbook(os.path.join(root, f))
            read_ws = read_wb.worksheets[0]
            for row in read_ws:
                for cell in row:
                    wb.active[cell.coordinate].value = cell.value
    del wb['Sheet']  # removes extra default sheet
    wb.save(res_path)

    Task.objects.filter(taskUUID=task.taskUUID).update(
        status="Completed")
    NodeGroup.objects.filter(currentTask=task.taskUUID).update(
        currentTask=None)
    return True


def checkNodes(nodes=LGNode.objects.all(), timeout=TIMEOUT_SECS, repeat=False):
    """
    Makes a health check to the provided clients, and updates status
    accordingly. If a node fails enough health-checks, it is
    removed from the database.
    :param nodes: A QuerySet or list-like object of clients to check.
    :param timeout: Time to respond before considering a node offline.
    :param repeat: Whether to recursively check nodes every 60s. Mainly used
    for the automatic checks done in the background on startup.
    :return:
    """
    import time
    start = time.time()
    availableNodes = []
    for node in nodes:
        nodeAddress = node.ip + ':' + str(node.port)
        try:
            print('checking: %s' % node)
            channel = grpc.insecure_channel(nodeAddress)
            nglm_pb2_grpc.ServerStub(channel)\
                .isAlive(nglm_pb2.query(query=''), timeout=timeout)
            print('checked: %s' % node)
            availableNodes.append(node)
            channel.close()
        except:
            uuidAttr = str(node.nodeUUID).replace('-', '_')
            if hasattr(checkNodes, uuidAttr):
                retries = getattr(checkNodes, uuidAttr)
                if retries >= 10:
                    LGNode.objects.filter(nodeUUID=node.nodeUUID).delete()
                    delattr(checkNodes, uuidAttr)
                else:
                    setattr(checkNodes, uuidAttr, retries + 1)
            else:
                setattr(checkNodes, uuidAttr, 1)
            continue
    end = time.time()
    print('Available nodes updated. Elapsed: %.2fs' % (end - start))
    updateNodes(availableNodes)
    if repeat:
        sleep(60)
        return checkNodes(repeat=True)
    return availableNodes


def scheduleTask(task):
    """
    Adds a task to the APScheduler BackgroundScheduler instance. Called on
    task creation.
    :param task: The Task model instance to schedule.
    :return: Returns an APScheduler job object.
    """
    job = SCHEDULER.add_job(
        startLogging, replace_existing=True,
        trigger='date', args=(task.assignedNode.nodes.all(), task),
        run_date=task.startTime, id=str(task.taskUUID)
    )
    if not SCHEDULER.running:
        SCHEDULER.start()
    print('Task with UUID %s scheduled.' % task.taskUUID)
    return job


def updateNodes(nodes):
    LGNode.objects.exclude(status="Busy").update(status="Offline")
    for node in nodes:
        LGNode.objects.exclude(status="Busy")\
            .filter(ip__iexact=node.ip).update(status="Available")


def saveResponse(chunks, path):
    """
    Saves a list-like object of byte chunks to a file.
    :param chunks: A list-like object of byte chunks.
    :param path: The destination path for the file.
    :return: None.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'wb') as f:
        for chunk in chunks:
            f.write(chunk.buffer)
        print('saved to %s' % path)


def startLogging(nodes, task):
    """
    Starts logging for a task, and updates the statuses accordingly.
    :param nodes: The nodes to begin the task for.
    :param task: The task to be executed.
    :return: None.
    """
    for node in nodes:
        nodeAddress = node.ip + ':' + str(node.port)
        try:
            channel = grpc.insecure_channel(nodeAddress)
            stub = nglm_pb2_grpc.LoggingStub(channel)
            itr = task.interval if task.interval else 2
            dur = int(task.duration.total_seconds()) if task.duration else 4
            params = nglm_pb2.params(pname='', interval=itr, duration=dur)

            response = stub.start(params)
            task.assignedNode.currentTask = task.taskUUID
            task.assignedNode.save()
            Task.objects.filter(taskUUID=task.taskUUID).update(
                status="In Progress")
            LGNode.objects.filter(nodeUUID=node.nodeUUID).update(
                status="Busy", currentTask=task.taskUUID)
        except Exception as e:
            Task.objects.filter(taskUUID=task.taskUUID).update(
                status="Failed: %s" % e)
            print(str(node) + ' was not available for task.')
            continue


def getChunks(f):
    while True:
        chunk = f.read(1024)
        if not chunk:
            break
        yield nglm_pb2.chunks(buffer=chunk)


def setConfig(nodes, f):
    """
    Sets the configuration file for the given clients, overriding the current.
    :param nodes: The nodes to set this configuration for.
    :param f: The File object of the configuration.
    :return: A list of clients in which setting failed. (Generally offline)
    """
    failedNodes = []
    chunkList = list(getChunks(f))
    for node in nodes:
        nodeAddress = node.ip + ':' + str(node.port)
        try:
            iterator = iter(chunkList)
            channel = grpc.insecure_channel(nodeAddress)
            stub = nglm_pb2_grpc.LoggingStub(channel)
            response = stub.setConfig(iterator)
            if not response.success:
                failedNodes.append(node.ip)
        except:
            failedNodes.append(node.ip)
            continue
    return failedNodes


def getConfig(nodes, size=1):
    """
    Retrieves the current configuration file(s) for the given client(s).
    :param nodes: The clients to retrieve the file from.
    :param size: The desired chunk size in KB. Defaults to 1KB.
    :return: Returns a list of clients that failed. (Generally offline)
    """
    failedNodes = []
    for node in nodes:
        try:
            nodeAddress = node.ip + ':' + str(node.port)
            channel = grpc.insecure_channel(nodeAddress)
            stub = nglm_pb2_grpc.LoggingStub(channel)
            params = nglm_pb2.chunkSize(size=size)

            response = stub.getConfig(params, timeout=TIMEOUT_SECS)
            saveResponse(response, os.path.join(
                os.path.dirname(sys.modules['__main__'].__file__),
                "nodeConfigs", str(node.nodeUUID) + '_config.ini'))
        except:
            failedNodes.append(node.ip)
            continue
    return failedNodes
